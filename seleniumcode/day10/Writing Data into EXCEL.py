import openpyxl
#Same data
# file="C:\\Users\\user\\OneDrive\\Desktop\\Book1.xlsx"
# workbook=openpyxl.load_workbook(file)
# sheet=workbook.active#opens only single sheet
# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(r,c).value="welcome"
#
# workbook.save(file)

file="C:\\Users\\user\\OneDrive\\Desktop\\Book1.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook.active
sheet.cell(1,1).value=123
sheet.cell(1,2).value="smith"
sheet.cell(1,3).value='john'

sheet.cell(2,1).value=456
sheet.cell(2,2).value="john"
sheet.cell(2,3).value='sam'

sheet.cell(3,1).value=123
sheet.cell(3,2).value="sak"
sheet.cell(3,3).value='dad'

workbook.save(file)