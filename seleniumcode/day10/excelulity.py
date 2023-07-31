import openpyxl
from openpyxl.styles import PatternFill

def getRowCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return (sheet.max_row)
def getColumnCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return (sheet.max_column)
def readData(file,SheetName,rownum,columnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[SheetName]
    return sheet.cell(rownum,columnno).value
def writeData(file, sheetNane, rownum, columnno, data):
    workbook =openpyxl.load_workbook(file)
    sheet =workbook[sheetNane]
    sheet.cell(rownum, columnno).value = data
    workbook.save(file)

def fillGreenColor(file,sheetName,rownum,columnno):
    workbook = openpyxl.load_workbook (file)
    sheet = workbook[sheetName]
    greenFill =PatternFill(start_color='60b212',
                          end_color='60b212',
                          fill_type='solid')
    sheet.cell(rownum,columnno).fill=greenFill
    workbook.save(file)

def fillRedColor(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    greenFill = PatternFill(start_color='ff0000',
                            end_color='ff0000',
                            fill_type='solid')
    sheet.cell(rownum, columnno).fill = greenFill
    workbook.save(file)


