import openpyxl
file="D:\\datascience\\shiva rama krishnna class\\9.DEEP LEARNIG\\Artifical neural network\\energy.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook['Sheet1']
rows=sheet.max_row
columns=sheet.max_column

#Reading all rows,Columns in Excel
for r in range(1,rows+1):
    for c in range(1,columns + 1):
        print(sheet.cell(r,c).value,end="   ")
    print()

